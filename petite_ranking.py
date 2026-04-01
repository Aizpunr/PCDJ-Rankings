import openpyxl, json, os, sys, re
from collections import defaultdict
sys.stdout.reconfigure(encoding='utf-8')

_dir = os.path.dirname(os.path.abspath(__file__))
_p = lambda f: os.path.join(_dir, f)

# Import CANONICAL from elo_75.py
elo_dir = os.path.join(os.path.dirname(_dir), 'zeepkist cotd elo')
sys.path.insert(0, elo_dir)
from elo_75 import CANONICAL

# Build name map from CANONICAL
NAME_MAP = {}
for canonical, aliases in CANONICAL.items():
    for alias in aliases:
        NAME_MAP[alias] = canonical

# Petite-specific aliases (names that appear differently in SGR's spreadsheet)
PETITE_ALIASES = {
    'Pandamane': 'PandaMane',
    'Pandamane ': 'PandaMane',
    'JakeAdjecent': 'JakeAdjacent',
    'JakeAdjecent ': 'JakeAdjacent',
    'Radsabsrad': 'RadAbsRad',
    'Streben': 'Sterben',
    'Mega Knight (Sterben)': 'Sterben',
    'Bowler (Sterben)': 'Sterben',
    'AndeMe17': 'AndMe',
    'Mackcheesy': 'MackCheesy',
    'An Actual G00se': 'An Actual g00se',
    'AndMe93': 'brrryy',
    'JustMaki': 'justMaki',
    'QuickRacer10': 'Quickracer10',
    'Redal': 'redal',
    '[GECK]R0nanC': 'R0nanC',
    '[CCC]Shinikage221': 'Shinikage221',
    '[Fae]Kyn': 'Kyn',
    'Brrry': 'brrryy',
    'Gilgool': 'gilgool',
    'BB_Benji': 'BB_Benji',
    'Ping': 'ping',
    'RoundNZT': 'RoundNzt',
    'Clowney': 'Clowny',
    'Brrryy': 'brrryy',
    'Magical': 'Magical',
    'Redstoney': 'Redstony',
    'r-tube': 'rtyyyyb',
    'rtube': 'rtyyyyb',
    'Mu': 'Mμ',
    'Lkat': 'LKat',
    'null/plexus': 'null/plexus',
    'Zoman': 'ZOMAN',
    'redal': 'redal',
    'A2 Zecklord': 'A2 Zecklord',
    'OLR94': 'SGR',
    'ShyGirlyRaccoon': 'SGR',
}
NAME_MAP.update(PETITE_ALIASES)

def strip_tag(name):
    return re.sub(r'\[.*?\]\s*', '', name).strip()

def normalize(name):
    name = name.strip()
    if name in NAME_MAP:
        return NAME_MAP[name]
    # Auto-strip: [TAG]Name -> Name if Name is a canonical key
    stripped = strip_tag(name)
    if stripped != name and (stripped in CANONICAL or stripped in NAME_MAP.values()):
        return NAME_MAP.get(stripped, stripped)
    return name

# --- Points system ---
POINTS_SCALE = 20  # multiplier: 1st=200, 2nd=180, ..., 10th=20
TROLL_MULT = 0.5   # troll/roulette at half points
BEST_OF_PCT = 0.70  # count best 70% of rounds per season

# Total events per season (from calendar/spreadsheet templates)
SEASON_TOTAL = {
    'Season 2': 17,   # 15 regular + Troll 2 + Roulette 2
    'Season 3': 22,   # 20 regular + Troll 3 + Roulette 3
}

def base_points(position):
    """10 for 1st, 9 for 2nd, ..., 1 for 10th."""
    if 1 <= position <= 10:
        return 11 - position
    return 0

def parse_petite(filepath):
    """Parse SGR's petite cup spreadsheet. Returns dict of season -> list of rounds."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    seasons = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        rounds = []

        for ri, row in enumerate(rows):
            # Scan each column group (4 columns wide, 6 groups per header row)
            for col_start in range(0, 26, 4):
                if col_start >= len(row) or row[col_start] is None:
                    continue
                header = str(row[col_start]).strip()
                if not (header.startswith('Round') or 'Troll' in header or 'Roulette' in header):
                    continue

                is_half = 'Troll' in header or 'Roulette' in header

                # Get lobby size from the "Players:" cell
                lobby_size = None
                if col_start + 2 < len(row) and row[col_start + 2] is not None:
                    try:
                        lobby_size = int(float(row[col_start + 2]))
                    except (ValueError, TypeError):
                        pass

                # Read the 10 player rows (2 rows below header)
                players = []
                for offset in range(2, 12):
                    if ri + offset >= len(rows):
                        break
                    data_row = rows[ri + offset]
                    if col_start >= len(data_row) or col_start + 1 >= len(data_row):
                        break
                    pos = data_row[col_start]
                    name = data_row[col_start + 1]
                    if name is None or pos is None:
                        continue
                    try:
                        pos = int(float(pos))
                    except (ValueError, TypeError):
                        continue
                    name = normalize(str(name))
                    pts = base_points(pos) * POINTS_SCALE
                    if is_half:
                        pts *= TROLL_MULT
                    players.append({
                        'position': pos,
                        'name': name,
                        'points': pts,
                    })

                if players:
                    rounds.append({
                        'name': header,
                        'lobby_size': lobby_size,
                        'is_half': is_half,
                        'players': players,
                    })

        if rounds:
            seasons[sheet_name] = rounds

    return seasons

def parse_cup_file(filepath, round_name, cup_label=None):
    """Parse a cup xlsx in COTD format. Finds cup by cup_label (e.g. 'Petite Cup 38').
    If cup_label is None, reads the first Position column found."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=False))

    # Find all Position headers and their associated cup names
    position_cols = []
    for ri, row in enumerate(rows):
        for cell in row:
            if cell.value == 'Position':
                # Look above for cup name
                cup_name = None
                for sr in range(cell.row - 2, max(0, cell.row - 5), -1):
                    val = ws.cell(row=sr, column=cell.column).value
                    if val and str(val).startswith('Petite'):
                        cup_name = str(val).strip()
                        break
                position_cols.append((cell.row, cell.column, cup_name))

    # Find the right cup
    target_col = None
    target_row = None
    for row_idx, col_idx, cname in position_cols:
        if cup_label and cname == cup_label:
            target_row, target_col = row_idx, col_idx
            break
        elif not cup_label and not target_col:
            target_row, target_col = row_idx, col_idx

    if target_col is None:
        print(f"  WARNING: Could not find cup '{cup_label}' in {filepath}")
        return {'name': round_name, 'lobby_size': 0, 'is_half': False, 'players': []}

    # Read players below the Position header
    players = []
    for r in range(target_row + 1, ws.max_row + 1):
        pos_val = ws.cell(row=r, column=target_col).value
        name_val = ws.cell(row=r, column=target_col + 1).value
        if pos_val is None or name_val is None:
            break
        try:
            pos = int(float(pos_val))
        except (ValueError, TypeError):
            continue
        name = normalize(str(name_val).strip())
        pts = base_points(pos) * POINTS_SCALE
        players.append({
            'position': pos,
            'name': name,
            'points': pts,
        })

    return {
        'name': round_name,
        'lobby_size': len(players),
        'is_half': False,
        'players': players,
    }

# Full-lobby replacements: override SGR's top-10-only data with complete results
# Format: 'Round N': ('filename.xlsx', 'Cup Label in xlsx')
FULL_LOBBY_REPLACEMENTS = {
    'Season 3': {
        'Round 5': ('Petite Cups 31-35.xlsx', 'Petite Cup 34'),
        'Round 6': ('Petite Cups 31-35.xlsx', 'Petite Cup 35'),
        'Round 8': ('Petite Cups 36-40.xlsx', 'Petite Cup 37'),
        'Round 9': ('Petite Cups 36-40.xlsx', 'Petite Cup 38'),
        'Round 10': ('Petite Cups 36-40.xlsx', 'Petite Cup 39'),
        'Troll 3': ('Petite Cups 36-40.xlsx', 'Petite Cup 40'),
    },
}

# Supplementary cups not yet in SGR's spreadsheet
EXTRA_CUPS = {
}

# --- Compute rankings ---
def compute_rankings(rounds, best_of, season_mode=True, championship=False):
    """Compute rankings from a list of rounds.

    season_mode=True: no drops until round count exceeds best_of (first 70% all count)
    season_mode=False: always use best_of (trailing/ATP mode)
    championship=True: OLR's 10-1 scale, no drops, troll/roulette excluded
    """
    player_rounds = defaultdict(list)

    for rnd in rounds:
        if championship and rnd['is_half']:
            continue  # troll/roulette give no points in championship mode
        for p in rnd['players']:
            pts = base_points(p['position']) if championship else p['points']
            player_rounds[p['name']].append({
                'round': rnd['name'],
                'points': pts,
                'position': p['position'],
                'is_half': rnd['is_half'],
            })

    rankings = []
    rounds_played_so_far = len(rounds)

    for name, entries in player_rounds.items():
        sorted_pts = sorted([e['points'] for e in entries], reverse=True)

        # Season mode: all results count until season passes 70% threshold
        if season_mode and rounds_played_so_far <= best_of:
            effective_best = len(sorted_pts)  # count everything
        else:
            effective_best = best_of

        best_n = sorted_pts[:effective_best]
        total_pts = sum(best_n)
        all_pts = sum(sorted_pts)

        wins = sum(1 for e in entries if e['position'] == 1)
        golds = sum(1 for e in entries if e['position'] == 1)
        silvers = sum(1 for e in entries if e['position'] == 2)
        bronzes = sum(1 for e in entries if e['position'] == 3)
        top5 = sum(1 for e in entries if e['position'] <= 5)
        top10 = len(entries)

        avg_pts = round(all_pts / top10, 1) if top10 > 0 else 0
        best_single = max(e['points'] for e in entries)
        best_pos = min(e['position'] for e in entries)
        avg_pos = round(sum(e['position'] for e in entries) / len(entries), 1)

        history = [{'r': e['round'], 'p': e['points'], 'pos': e['position']} for e in entries]

        rankings.append({
            'name': name,
            'points': total_pts,
            'points_all': all_pts,
            'rounds': top10,
            'wins': wins,
            'podiums': {'gold': golds, 'silver': silvers, 'bronze': bronzes},
            'top5': top5,
            'avg_pts': avg_pts,
            'avg_pos': avg_pos,
            'best_single': best_single,
            'best_pos': best_pos,
            'dropped': max(0, len(sorted_pts) - effective_best),
            'history': history,
        })

    rankings.sort(key=lambda p: (p['points'], p['wins'], p['podiums']['gold'] + p['podiums']['silver'] + p['podiums']['bronze'], p['avg_pts']), reverse=True)
    for i, p in enumerate(rankings, 1):
        p['rank'] = i

    return rankings

# --- Cup Strength (SOF) ---
# Uses COTD weighted ELO to measure petite lobby quality
# PCDJ cup N ≈ COTD cup (97 + N) (both weekly events)
# Strength = avg normalized ELO of 10 highest-rated players IN THE LOBBY (by ELO, not by finish position)
PCDJ_TO_COTD = 97
POOL_CAP = 196

def load_cotd_histories():
    """Load COTD player histories for per-cup ELO lookups."""
    try:
        with open(os.path.join(elo_dir, 'alldata.json'), encoding='utf-8') as f:
            cotd = json.load(f)
        histories = {}
        for p in cotd['weighted']:
            histories[p['n']] = p['h']
        return histories
    except FileNotFoundError:
        print("  WARNING: alldata.json not found, cup strength unavailable")
        return None

def get_elo_at_cup(histories, name, cotd_cup):
    """Get player's ELO just before a given COTD cup number."""
    if name not in histories:
        return None
    hist = histories[name]
    before = [h for h in hist if h['c'] < cotd_cup]
    if before:
        return before[-1]['r']
    if hist and hist[0]['c'] == cotd_cup:
        return 1500
    return None

def compute_cup_strength(rnd, pcdj_cup, histories):
    """Compute strength for a petite round using COTD ELO.
    Finds the 10 highest-rated players IN THE LOBBY (by ELO, not finish position),
    normalizes their ratings, and averages them."""
    if not histories or len(rnd['players']) <= 10:
        return None
    cotd_cup = PCDJ_TO_COTD + pcdj_cup

    # Build normalized pool snapshot at this COTD cup
    pool = {}
    for name in histories:
        elo = get_elo_at_cup(histories, name, cotd_cup)
        if elo:
            pool[name] = elo
    pool_sorted = sorted(pool.items(), key=lambda x: x[1], reverse=True)[:POOL_CAP]
    if not pool_sorted:
        return None
    max_r = pool_sorted[0][1]
    sc = 2000 / max_r
    norm = {n: r * sc for n, r in pool_sorted}
    min_pool = min(norm.values())

    # Find 10 highest ELO players IN THE LOBBY (by ELO, NOT by finish position)
    lobby_elos = []
    for p in rnd['players']:
        if p['name'] in norm:
            lobby_elos.append(norm[p['name']])
    lobby_elos.sort(reverse=True)
    top10 = lobby_elos[:10]
    while len(top10) < 10:
        top10.append(min_pool)
    avg = sum(top10) / 10
    return round(avg / 1850 * 100, 1)

# Season round number -> overall PCDJ cup number
SEASON_CUP_OFFSET = {
    'Season 2': 0,   # S2 Round 1 = Cup 1
    'Season 3': 29,  # S3 Round 1 = Cup 30
}

# --- Main ---
print("Parsing petite cup data...")
seasons = parse_petite(_p('Results Petite .xlsx'))

# Replace rounds with full-lobby data where available
for season_name, replacements in FULL_LOBBY_REPLACEMENTS.items():
    if season_name not in seasons:
        continue
    for round_name, (filename, cup_label) in replacements.items():
        rnd = parse_cup_file(_p(filename), round_name, cup_label=cup_label)
        rnd['is_half'] = 'Troll' in round_name or 'Roulette' in round_name
        if rnd['is_half']:
            for p in rnd['players']:
                p['points'] *= TROLL_MULT
        if not rnd['players']:
            continue
        # Find and replace the existing round
        replaced = False
        for i, existing in enumerate(seasons[season_name]):
            if existing['name'] == round_name:
                seasons[season_name][i] = rnd
                replaced = True
                print(f"  Replaced {round_name} with full lobby from {filename}: {len(rnd['players'])} players")
                break
        if not replaced:
            seasons[season_name].append(rnd)
            print(f"  Added {round_name} from {filename}: {len(rnd['players'])} players")

# Append supplementary cups
for season_name, cups in EXTRA_CUPS.items():
    if season_name not in seasons:
        seasons[season_name] = []
    for filename, round_name in cups:
        rnd = parse_cup_file(_p(filename), round_name)
        if rnd['players']:
            seasons[season_name].append(rnd)
            print(f"  Added {round_name} from {filename}: {len(rnd['players'])} top-10 players, {rnd['lobby_size']} total")
season_names = list(seasons.keys())

output = {}

# --- Season rankings ---
for season_name, rounds in seasons.items():
    regular = [r for r in rounds if not r['is_half']]
    special = [r for r in rounds if r['is_half']]
    total_events = SEASON_TOTAL.get(season_name, len(rounds))
    best_of = round(total_events * BEST_OF_PCT)
    rounds_played = len(rounds)
    drops_active = rounds_played > best_of

    print(f"\n=== {season_name} ===")
    print(f"  {len(regular)} regular + {len(special)} special (half pts)")
    print(f"  {rounds_played}/{total_events} rounds played, best {best_of} count (70% of {total_events})")
    if drops_active:
        print(f"  Drops active — dropping {rounds_played - best_of} worst")
    else:
        print(f"  All results count (drops start at round {best_of + 1})")

    # Compute cup strength for full-lobby rounds
    cotd_histories = load_cotd_histories()
    round_strengths = {}
    offset = SEASON_CUP_OFFSET.get(season_name, 0)
    for rnd in rounds:
        rnum = int(''.join(c for c in rnd['name'] if c.isdigit()) or 0)
        pcdj_cup = offset + rnum
        sof = compute_cup_strength(rnd, pcdj_cup, cotd_histories)
        if sof is not None:
            round_strengths[rnd['name']] = sof
            print(f"  {rnd['name']} (PCDJ {pcdj_cup}): SOF {sof}%")

    rankings = compute_rankings(rounds, best_of, season_mode=True)
    # Championship points: OLR's original 10-1 scale, no drops, troll/roulette excluded
    champ_rankings = compute_rankings(rounds, len(rounds), season_mode=False, championship=True)
    output[season_name] = {
        'type': 'season',
        'total_events': total_events,
        'rounds_played': rounds_played,
        'rounds_regular': len(regular),
        'rounds_special': len(special),
        'best_of': best_of,
        'drops_active': drops_active,
        'points_scale': POINTS_SCALE,
        'rankings': rankings,
        'championship': champ_rankings,
        'round_strengths': round_strengths,
    }

    print(f"\n{'#':<4}{'Player':<22}{'Pts':<8}{'Rnds':<6}{'W':<4}{'Pod':<5}{'T5':<4}{'Avg Pts':<9}{'Avg Pos':<9}{'Drop'}")
    print("=" * 75)
    for p in rankings[:20]:
        print(f"{p['rank']:<4}{p['name']:<22}{p['points']:<8.0f}{p['rounds']:<6}{p['wins']:<4}{p['podiums']['gold']}/{p['podiums']['silver']}/{p['podiums']['bronze']:<4}{p['avg_pts']:<9}{p['avg_pos']:<9}{p['dropped']}")

# --- Trailing ranking (last 2 seasons combined, ATP-style) ---
if len(season_names) >= 2:
    trailing_seasons = season_names[-2:]
    trailing_rounds = []
    for sn in trailing_seasons:
        # Prefix round names with season to avoid collisions (both have "Round 1", etc.)
        for rnd in seasons[sn]:
            prefixed = dict(rnd)
            prefixed['name'] = f"{sn} {rnd['name']}"
            prefixed['players'] = list(rnd['players'])  # shallow copy
            trailing_rounds.append(prefixed)
    best_of_trailing = round(len(trailing_rounds) * BEST_OF_PCT)

    print(f"\n=== PCDJ Ranking ({' + '.join(trailing_seasons)}) ===")
    print(f"  {len(trailing_rounds)} total rounds, best {best_of_trailing} (70%)")

    trailing_rankings = compute_rankings(trailing_rounds, best_of_trailing, season_mode=False)
    trailing_champ = compute_rankings(trailing_rounds, len(trailing_rounds), season_mode=False, championship=True)
    output['PCDJ Ranking'] = {
        'type': 'trailing',
        'seasons': trailing_seasons,
        'rounds_total': len(trailing_rounds),
        'best_of': best_of_trailing,
        'points_scale': POINTS_SCALE,
        'rankings': trailing_rankings,
        'championship': trailing_champ,
        'round_strengths': {},  # trailing combines seasons, strengths are per-season
    }

    print(f"\n{'#':<4}{'Player':<22}{'Pts':<8}{'Rnds':<6}{'W':<4}{'Pod':<5}{'T5':<4}{'Avg Pts':<9}{'Avg Pos':<9}{'Drop'}")
    print("=" * 75)
    for p in trailing_rankings[:20]:
        print(f"{p['rank']:<4}{p['name']:<22}{p['points']:<8.0f}{p['rounds']:<6}{p['wins']:<4}{p['podiums']['gold']}/{p['podiums']['silver']}/{p['podiums']['bronze']:<4}{p['avg_pts']:<9}{p['avg_pos']:<9}{p['dropped']}")

# Save JSON
with open(_p('petite_rankings.json'), 'w') as f:
    json.dump(output, f, indent=2)
print(f"\npetite_rankings.json saved")
