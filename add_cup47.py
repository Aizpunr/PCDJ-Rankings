"""Add Cup 47 to Petite Cups 46-50.xlsx in cols 7-10.

Source: cup logs/petite_47_reconstructed.json
Excluded: SGR (tested the map, both mappers otherwise not in lobby per user 2026-05-20)
Maps: TBD (fill row 3 manually from cup announcement).
"""
import json
import os
import openpyxl

base = os.path.dirname(os.path.abspath(__file__))
xlsx_path = os.path.join(base, 'Petite Cups 46-50.xlsx')
json_path = os.path.join(base, 'cup logs', 'petite_47_reconstructed.json')

EXCLUDE = {'[CSC]ShyGirlyRaccoon', 'ShyGirlyRaccoon', 'SGR'}

with open(json_path, 'r', encoding='utf-8') as f:
    raw = json.load(f)

filtered = [r for r in raw if r['name'] not in EXCLUDE]

# Re-rank preserving tie groups (same as add_cup46 pattern)
new_list = []
prev_pos = None
new_pos = 0
seen = 0
for entry in filtered:
    seen += 1
    if entry['pos'] != prev_pos:
        new_pos = seen
        prev_pos = entry['pos']
    new_list.append({**entry, 'new_pos': new_pos})

print(f'Total entries after mapper exclusion: {len(new_list)}')
for e in new_list:
    t = f"{e['time']:.5f}" if e['time'] is not None else 'DNF'
    safe = e['name'].encode('ascii', 'replace').decode('ascii')
    print(f"  {e['new_pos']:3} {safe:38} {t:>10}  (R{e['round'] or '-'})")

wb = openpyxl.load_workbook(xlsx_path)
ws = wb.active

COL = 7  # cup 46 in cols 1-4 (+2 spacer); cup 47 starts at col 7
ws.cell(row=2, column=COL, value='Petite Cup 47')
ws.cell(row=3, column=COL, value='Maps: <fill from cup announcement> + <map 2>')
ws.cell(row=5, column=COL,   value='Position')
ws.cell(row=5, column=COL+1, value='Name')
ws.cell(row=5, column=COL+2, value='Elim Time')
ws.cell(row=5, column=COL+3, value='Elim Round')

# Clear any old data in cup 47 area (rows 6+)
for r in range(6, 60):
    for c in range(COL, COL+4):
        ws.cell(row=r, column=c, value=None)

for i, e in enumerate(new_list):
    r = 6 + i
    ws.cell(row=r, column=COL,   value=e['new_pos'])
    ws.cell(row=r, column=COL+1, value=e['name'])
    if e['time'] is not None:
        ws.cell(row=r, column=COL+2, value=round(e['time'], 5))
    else:
        ws.cell(row=r, column=COL+2, value='DNF')
    if e['round'] is not None and e.get('note') != 'winner':
        ws.cell(row=r, column=COL+3, value=e['round'])

wb.save(xlsx_path)
print(f'\nWrote Cup 47 -> {xlsx_path}')
